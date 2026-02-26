#!/usr/bin/env python3
"""
한 달 식단 짜기 - 프로토타입
카테고리: 주식, 국/스프, 반찬, 기타 (JSON 관리, 저장/수정 가능)
레이아웃: 좌측(6) 달력, 우측(4) 메뉴 선택. data.json 통합 저장/로드.
"""

import json
import sys
from calendar import monthrange
from datetime import datetime
from pathlib import Path
from tkinter import messagebox

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None

try:
    import customtkinter as ctk
except ImportError as e:
    err = str(e).lower()
    if "_tkinter" in err or "no module named" in err:
        print("이 Python에는 GUI용 tkinter가 없습니다.")
        print("macOS Homebrew 사용 시: brew install python-tk@3.13")
        print("또는 시스템 Python(/usr/bin/python3)으로 실행해 보세요.")
    else:
        print("customtkinter가 필요합니다. 터미널에서 실행하세요:")
        print("  python3 -m venv .venv && .venv/bin/pip install -r requirements.txt")
        print("  .venv/bin/python main.py")
    sys.exit(1)

# 데이터 저장 경로 (통합 JSON: 앱 실행 시 여기서 모두 불러옴)
# exe/앱으로 빌드된 경우 실행 파일이 있는 폴더에 data.json·export 저장
if getattr(sys, "frozen", False):
    DATA_DIR = Path(sys.executable).resolve().parent
else:
    DATA_DIR = Path(__file__).resolve().parent
EXPORT_DIR = DATA_DIR / "export"  # 엑셀 다운로드 저장 경로
DATA_FILE = DATA_DIR / "data.json"  # 메뉴 + 식단 계획 통합
PLAN_FILE = DATA_DIR / "meal_plan.json"  # 하위 호환
MENUS_FILE = DATA_DIR / "menus.json"  # 하위 호환

# 카테고리 기본값 (menus.json 없을 때 사용)
DEFAULT_MENUS = {
    "주식": ["밥", "현미밥", "잡곡밥", "죽", "라면", "짜장면", "칼국수", "비빔밥", "덮밥", "국밥"],
    "국/스프": ["미역국", "된장국", "김치찌개", "된장찌개", "순두부찌개", "배추국", "콩나물국", "우동", "만두국", "스프"],
    "반찬": ["김치", "나물", "계란말이", "제육볶음", "멸치볶음", "감자조림", "두부조림", "시금치나물", "깻잎지", "오이무침"],
    "기타": ["샐러드", "과일", "유제품", "과자", "떡", "김밥", "삼각김밥", "도시락", "외식", "기타"],
}

CATEGORIES = ["주식", "국/스프", "반찬", "기타"]


def _default_slots():
    return {c: 1 for c in CATEGORIES}


def load_all():
    """
    통합 data.json에서 메뉴+식단 계획+날짜별 슬롯 수 로드.
    day_slots[month][day_str] = {"주식": 1, "국/스프": 2, ...} 로 날짜마다 독립.
    """
    if DATA_FILE.exists():
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            menus = data.get("menus")
            plans = data.get("plans", {})
            day_slots = data.get("day_slots", {})
            if not isinstance(day_slots, dict):
                day_slots = {}
            # 예전 slots_per_category(전역) 있으면 무시, 날짜별은 비어 있으면 기본 1행
            if menus:
                for c in CATEGORIES:
                    if c not in menus or not menus[c]:
                        menus[c] = DEFAULT_MENUS[c].copy()
                return menus, plans, day_slots
        except (json.JSONDecodeError, IOError):
            pass
    menus = load_menus_legacy()
    plan_data = load_plan_legacy()
    plans = plan_data.get("plans", {})
    return menus, plans, {}


def load_menus_legacy():
    """menus.json 또는 기본 메뉴"""
    if MENUS_FILE.exists():
        try:
            with open(MENUS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            for cat in CATEGORIES:
                if cat not in data or not data[cat]:
                    data[cat] = DEFAULT_MENUS[cat].copy()
            return data
        except (json.JSONDecodeError, IOError):
            pass
    return {c: list(items) for c, items in DEFAULT_MENUS.items()}


def load_plan_legacy():
    """meal_plan.json 또는 빈 계획"""
    if PLAN_FILE.exists():
        try:
            with open(PLAN_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {"plans": {}}


def save_all(menus, plans, day_slots=None):
    """메뉴 + 식단 계획 + 날짜별 슬롯 수를 통합 data.json에 저장"""
    if day_slots is None:
        day_slots = {}
    data = {"menus": menus, "plans": plans, "day_slots": day_slots}
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


class MenuEditWindow(ctk.CTkToplevel):
    """메뉴 편집 창 (카테고리별 추가/삭제, JSON 저장). 목록은 선택만 가능, 삭제는 버튼으로."""
    def __init__(self, parent, menus, on_save):
        super().__init__(parent)
        self.categories = list(CATEGORIES)
        self.menus = {c: list(menus.get(c, [])) for c in self.categories}
        for c in self.categories:
            if c not in self.menus:
                self.menus[c] = []
        self.on_save = on_save
        self.selected_entries = {c: None for c in self.categories}
        self.title("메뉴 편집")
        self.geometry("450x420")
        self.transient(parent)

        self._build_ui()

    def _build_ui(self):
        self.tabview = ctk.CTkTabview(self, width=400)
        self.tabview.pack(fill="both", expand=True, padx=15, pady=15)
        self.list_frames = {}
        self.entries = {}
        self.delete_buttons = {}

        for cat in self.categories:
            tab = self.tabview.add(cat)
            ctk.CTkLabel(tab, text=f"'{cat}' 메뉴 목록", font=ctk.CTkFont(weight="bold")).pack(anchor="w", pady=(0, 5))
            list_frame = ctk.CTkScrollableFrame(tab, height=200)
            list_frame.pack(fill="both", expand=True)
            self.list_frames[cat] = list_frame

            add_f = ctk.CTkFrame(tab, fg_color="transparent")
            add_f.pack(fill="x", pady=8)
            ent = ctk.CTkEntry(add_f, width=180, placeholder_text="새 메뉴 이름")
            ent.pack(side="left", padx=(0, 8))
            self.entries[cat] = ent
            ctk.CTkButton(add_f, text="추가", width=60, command=lambda c=cat: self._add(c)).pack(side="left", padx=2)
            self.delete_buttons[cat] = ctk.CTkButton(add_f, text="삭제", width=60, command=lambda c=cat: self._delete(c), state="disabled", fg_color="gray50")
            self.delete_buttons[cat].pack(side="left", padx=2)

        btn_f = ctk.CTkFrame(self, fg_color="transparent")
        btn_f.pack(fill="x", padx=15, pady=(0, 15))
        ctk.CTkButton(btn_f, text="저장", fg_color="green", hover_color="darkgreen", command=self._save).pack(side="right", padx=5)
        ctk.CTkButton(btn_f, text="취소", command=self.destroy).pack(side="right")

        for cat in self.categories:
            self._refresh_display(cat)

    def _select(self, cat, name):
        if self.selected_entries[cat] == name:
            self.selected_entries[cat] = None
        else:
            self.selected_entries[cat] = name
        self._refresh_display(cat)
        self.delete_buttons[cat].configure(state="normal" if self.selected_entries[cat] else "disabled", fg_color=("gray50", "gray50") if not self.selected_entries[cat] else ("#c94c4c", "#8b3535"))

    def _add(self, cat):
        name = self.entries[cat].get().strip()
        if not name:
            return
        self.entries[cat].delete(0, "end")  # 한글 IME 잔여 글자 방지: 먼저 비움
        if cat not in self.menus:
            self.menus[cat] = []
        if name in self.menus[cat]:
            return
        self.menus[cat].append(name)
        self.selected_entries[cat] = name  # 추가한 메뉴로 선택 이동
        self._refresh_display(cat)
        self._scroll_to_bottom(cat)
        self.entries[cat].focus_set()
        # 포커스 복귀 후 IME가 마지막 글자를 다시 넣는 경우 대비, 잠시 뒤 한 번 더 비움
        self.after(100, lambda: self._clear_entry(cat))

    def _clear_entry(self, cat):
        if cat in self.entries:
            self.entries[cat].delete(0, "end")

    def _refresh_display(self, cat):
        """목록을 버튼으로 채움 (직접 수정 불가, 클릭 시 선택)"""
        frame = self.list_frames[cat]
        for w in frame.winfo_children():
            w.destroy()
        for name in self.menus.get(cat, []):
            is_selected = self.selected_entries.get(cat) == name
            btn = ctk.CTkButton(
                frame,
                text=name,
                anchor="w",
                fg_color=("#3b8ed0", "#1f6aa5") if is_selected else ("gray85", "gray25"),
                text_color=("black", "white"),
                command=lambda c=cat, n=name: self._select(c, n),
            )
            btn.pack(fill="x", pady=2)
        self.delete_buttons[cat].configure(state="normal" if self.selected_entries.get(cat) else "disabled", fg_color=("gray50", "gray50") if not self.selected_entries.get(cat) else ("#c94c4c", "#8b3535"))

    def _scroll_to_bottom(self, cat):
        """스크롤을 맨 아래로 이동해 새로 추가된 메뉴가 보이도록"""
        frame = self.list_frames[cat]
        frame.update_idletasks()
        canvas = frame.master
        if hasattr(canvas, "yview_moveto"):
            canvas.yview_moveto(1.0)

    def _delete(self, cat):
        name = self.selected_entries.get(cat)
        if not name or name not in self.menus.get(cat, []):
            return
        self.menus[cat].remove(name)
        self.selected_entries[cat] = None
        self._refresh_display(cat)

    def _save(self):
        for c in self.categories:
            if not self.menus.get(c):
                self.menus[c] = DEFAULT_MENUS[c].copy() if c in CATEGORIES else []
        self.on_save(self.menus)
        self.destroy()


class MealPlannerApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("한 달 식단 짜기")
        self.geometry("1000x620")
        self.minsize(900, 550)

        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.menus, self.plans, self.day_slots = load_all()
        self.slots_per_category = _default_slots().copy()  # 현재 선택한 날짜의 행 개수

        self.current_year = datetime.now().year
        self.current_month = datetime.now().month
        self.selected_day = None
        self.day_buttons = {}

        self._build_ui()

    def _build_ui(self):
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=15)
        self.main_frame.grid_columnconfigure(0, weight=1)   # 처음엔 달력만 → 좌측만 확장
        self.main_frame.grid_columnconfigure(1, weight=0)   # 우측 숨김 시 0

        # ----- 좌측: 월 선택 + 달력 -----
        left_panel = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        left_panel.grid_columnconfigure(0, weight=1)
        left_panel.grid_rowconfigure(2, weight=1)

        top_left = ctk.CTkFrame(left_panel, fg_color="transparent")
        top_left.grid(row=0, column=0, sticky="ew", pady=(0, 4))
        top_left.grid_columnconfigure(1, weight=1)
        ctk.CTkButton(top_left, text="◀ 이전", width=80, command=self._prev_month).grid(row=0, column=0, padx=(0, 10))
        self.month_label = ctk.CTkLabel(top_left, text="", font=ctk.CTkFont(size=18, weight="bold"))
        self.month_label.grid(row=0, column=1)
        ctk.CTkButton(top_left, text="다음 ▶", width=80, command=self._next_month).grid(row=0, column=2, padx=(10, 0))

        export_row = ctk.CTkFrame(left_panel, fg_color="transparent")
        export_row.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        export_row.grid_columnconfigure(0, weight=1)
        ctk.CTkButton(export_row, text="📥 엑셀 다운로드", width=120, command=self._export_excel, fg_color="green", hover_color="darkgreen").pack(side="right")

        self.calendar_frame = ctk.CTkFrame(left_panel, fg_color="transparent")
        self.calendar_frame.grid(row=2, column=0, sticky="nsew")
        for col, w in enumerate(["일", "월", "화", "수", "목", "금", "토"]):
            ctk.CTkLabel(self.calendar_frame, text=w, font=ctk.CTkFont(weight="bold")).grid(row=0, column=col, padx=2, pady=2, sticky="ew")
        self.calendar_frame.columnconfigure(tuple(range(7)), weight=1)

        # ----- 우측: 메뉴 선택 (날짜 클릭 시에만 표시) -----
        self.right_panel = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        self.right_panel.grid(row=0, column=1, sticky="nsew")
        self.right_panel.grid_columnconfigure(0, weight=1)
        self.right_panel.grid_remove()   # 처음엔 숨김

        ctk.CTkLabel(self.right_panel, text="선택한 날짜", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        self.selected_label = ctk.CTkLabel(self.right_panel, text="날짜를 클릭하세요", text_color="gray")
        self.selected_label.pack(anchor="w", pady=(0, 12))

        self.categories_frame = ctk.CTkFrame(self.right_panel, fg_color="transparent")
        self.categories_frame.pack(fill="both", expand=True)
        self.category_vars = {}   # (cat, slot) -> StringVar
        self.category_combos = {}  # (cat, slot) -> ComboBox
        self._build_category_rows()

        btn_row = ctk.CTkFrame(self.right_panel, fg_color="transparent")
        btn_row.pack(fill="x", pady=8)
        ctk.CTkButton(btn_row, text="✏️ 메뉴 편집", command=self._open_menu_edit, fg_color="gray40").pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="저장", command=self._save, fg_color="green", hover_color="darkgreen").pack(side="left")

        self._refresh_month()

    def _build_category_rows(self):
        """카테고리별 슬롯 수만큼 행 생성: 주식 [select] [+] / 주식 [select] [-] ..."""
        for w in self.categories_frame.winfo_children():
            w.destroy()
        self.category_vars.clear()
        self.category_combos.clear()
        for cat in CATEGORIES:
            for slot in range(self.slots_per_category.get(cat, 1)):
                row = ctk.CTkFrame(self.categories_frame, fg_color="transparent")
                row.pack(fill="x", pady=4)
                ctk.CTkLabel(row, text=f"{cat}:", width=70, anchor="w").pack(side="left", padx=(0, 8))
                var = ctk.StringVar(value="")
                values = [""] + self.menus.get(cat, [])
                combo = ctk.CTkComboBox(row, values=values, variable=var, width=200)
                combo.pack(side="left", padx=(0, 8))
                key = (cat, slot)
                self.category_vars[key] = var
                self.category_combos[key] = combo
                if slot == 0:
                    ctk.CTkButton(row, text="+", width=36, command=lambda c=cat: self._add_slot(c), fg_color="green", hover_color="darkgreen").pack(side="left")
                else:
                    ctk.CTkButton(row, text="-", width=36, command=lambda c=cat: self._remove_slot(c), fg_color="#c94c4c", hover_color="#8b3535").pack(side="left")

    def _add_slot(self, cat):
        """해당 카테고리 선택 행 한 줄 추가. 현재 선택 내용 유지."""
        self._apply_meal()  # 먼저 현재 선택을 plan에 반영
        self.slots_per_category[cat] = self.slots_per_category.get(cat, 1) + 1
        self._save_day_slots()
        self._build_category_rows()
        self._repopulate_rows_from_plan()

    def _remove_slot(self, cat):
        """해당 카테고리의 추가 행 한 줄 제거 (최소 1행 유지). 현재 선택 내용 유지."""
        if self.slots_per_category.get(cat, 1) <= 1:
            return
        self._apply_meal()
        self.slots_per_category[cat] -= 1
        self._save_day_slots()
        self._build_category_rows()
        self._repopulate_rows_from_plan()

    def _save_day_slots(self):
        """현재 선택한 날짜의 슬롯 구성을 day_slots에 저장"""
        if self.selected_day is None:
            return
        key = self._month_key()
        if key not in self.day_slots:
            self.day_slots[key] = {}
        self.day_slots[key][str(self.selected_day)] = {c: self.slots_per_category.get(c, 1) for c in CATEGORIES}

    def _repopulate_rows_from_plan(self):
        """선택한 날짜가 있으면 저장된 계획을 현재 행 변수에 다시 채움 (행 추가/삭제 후 바인딩 유지)"""
        if self.selected_day is None:
            return
        key = self._month_key()
        plan = self.plans.get(key, {})
        current = plan.get(str(self.selected_day), "")
        if not current:
            return
        raw = [p.strip() for p in current.split(" | ")]
        n = len(CATEGORIES)
        parts = ([""] + raw + [""] * max(0, n - len(raw) - 1))[:n] if len(raw) < n else raw[:n]
        for i, cat in enumerate(CATEGORIES):
            slot_vals = [s.strip() for s in parts[i].strip().split(",") if s.strip()] or [""]
            for slot in range(self.slots_per_category.get(cat, 1)):
                k = (cat, slot)
                if k in self.category_vars:
                    self.category_vars[k].set(slot_vals[slot] if slot < len(slot_vals) else "")

    def _show_right_panel(self):
        self.main_frame.grid_columnconfigure(0, weight=4)
        self.main_frame.grid_columnconfigure(1, weight=6)
        self.right_panel.grid()

    def _hide_right_panel(self):
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_columnconfigure(1, weight=0)
        self.right_panel.grid_remove()
        self.selected_day = None
        self.selected_label.configure(text="날짜를 클릭하세요", text_color="gray")
        for d, (btn, _) in self.day_buttons.items():
            btn.configure(border_width=0, border_color=("gray85", "gray25"))

    def _export_excel(self):
        if Workbook is None:
            messagebox.showinfo("알림", "openpyxl이 필요합니다.\npip install openpyxl")
            return
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Border, Side, PatternFill
        key = self._month_key()
        plan = self.plans.get(key, {})
        first_weekday, num_days = monthrange(self.current_year, self.current_month)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        day_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        def plan_to_items(line):
            """'주식 | 국/스프 | 반찬 | 기타' 형식을 메뉴 항목 리스트로 변환 (쉼표 분리 포함)."""
            if not line or not line.strip():
                return []
            items = []
            for part in line.split(" | "):
                for x in part.split(","):
                    x = x.strip()
                    if x:
                        items.append(x)
            return items

        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.current_year}-{self.current_month:02d}"
        ncols = 7
        # 상단 중간: "2026년 2월" (셀 병합 없이 4번째 열에만 표기, 1행 전체 테두리)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.border = border
        title_cell = ws.cell(row=1, column=4, value=f"{self.current_year}년 {self.current_month}월")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        # 요일 헤더: 일 | 월 | 화 | 수 | 목 | 금 | 토
        week_headers = ["일", "월", "화", "수", "목", "금", "토"]
        for c, h in enumerate(week_headers, 1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        # 달력 그리드: 주마다 최소 4행 (날짜 1행 + 메뉴 행은 각 메뉴당 1셀, 최소 3행)
        num_weeks = max(1, (first_weekday + num_days + 6) // 7)
        current_row = 3
        for week_row in range(num_weeks):
            # 이번 주 7일의 메뉴 리스트 수집
            week_days = []
            for col_0 in range(7):
                pos = week_row * 7 + col_0
                day_num = (pos - first_weekday + 1) if (pos >= first_weekday and pos < first_weekday + num_days) else None
                day_str = str(day_num) if day_num else ""
                items = plan_to_items(plan.get(day_str, "")) if day_str else []
                week_days.append((day_num, items))
            menu_rows = max(3, max(len(items) for _, items in week_days))
            base_row = current_row
            current_row = base_row + 1 + menu_rows
            # 날짜 행 (일자에 색상 + 테두리)
            for col_0 in range(7):
                day_num, _ = week_days[col_0]
                cell = ws.cell(row=base_row, column=col_0 + 1, value=day_num if day_num is not None else "")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                cell.fill = day_fill
            # 메뉴 행: 각 메뉴당 각자 셀 (병합 없음)
            for r in range(menu_rows):
                for col_0 in range(7):
                    _, items = week_days[col_0]
                    val = items[r] if r < len(items) else ""
                    cell = ws.cell(row=base_row + 1 + r, column=col_0 + 1, value=val)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = border
        # 주 간 행 수가 가변이므로 다음 주 base_row 계산을 위해 rows_per_week 사용하지 않음 (이미 위에서 1+menu_rows로 처리)
        # 열 너비
        for c in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 14
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = EXPORT_DIR / f"식단_{self.current_year}-{self.current_month:02d}.xlsx"
        try:
            wb.save(out_path)
            messagebox.showinfo("알림", f"저장되었습니다.\n{out_path.name}\n\n경로: {out_path}")
        except Exception as e:
            messagebox.showerror("저장 실패", str(e))

    def _open_menu_edit(self):
        def on_save(menus):
            self.menus = menus
            for key in self.category_combos:
                cat = key[0]
                self.category_combos[key].configure(values=[""] + self.menus.get(cat, []))
            save_all(self.menus, self.plans, self.day_slots)

        win = MenuEditWindow(self, self.menus, on_save)
        win.focus_set()
        win.grab_set()

    def _month_key(self):
        return f"{self.current_year}-{self.current_month:02d}"

    def _prev_month(self):
        if self.current_month == 1:
            self.current_month = 12
            self.current_year -= 1
        else:
            self.current_month -= 1
        self._refresh_month()

    def _next_month(self):
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self._refresh_month()

    def _refresh_month(self):
        self.month_label.configure(text=f"{self.current_year}년 {self.current_month}월")
        for w in self.calendar_frame.winfo_children():
            w.destroy()
        self.day_buttons.clear()
        key = self._month_key()
        plan = self.plans.get(key, {})
        first_weekday, num_days = monthrange(self.current_year, self.current_month)

        for col, w in enumerate(["일", "월", "화", "수", "목", "금", "토"]):
            ctk.CTkLabel(self.calendar_frame, text=w, font=ctk.CTkFont(weight="bold")).grid(row=0, column=col, padx=2, pady=2, sticky="ew")

        for day in range(1, num_days + 1):
            day_str = str(day)
            meal_text = plan.get(day_str, "")
            if meal_text:
                meal_display = meal_text.replace(" | ", "\n").replace(",", "\n")
            else:
                meal_display = ""
            pos = first_weekday + (day - 1)
            row, col = 1 + pos // 7, pos % 7
            btn = ctk.CTkButton(
                self.calendar_frame,
                text=f"{day}\n{meal_display}" if meal_display else str(day),
                height=62,
                font=ctk.CTkFont(size=12),
                anchor="n",
                command=lambda d=day: self._select_day(d),
                fg_color=("gray85", "gray25") if not meal_text else ("#3b8ed0", "#1f6aa5"),
                text_color=("gray10", "gray90"),
            )
            btn.grid(row=row, column=col, padx=2, pady=2, sticky="nsew")
            self.day_buttons[day] = (btn, meal_text)

        for r in range(1, 7):
            self.calendar_frame.rowconfigure(r, weight=1)
        self.calendar_frame.columnconfigure(tuple(range(7)), weight=1)

        self.selected_day = None
        self.selected_label.configure(text="날짜를 클릭하세요", text_color="gray")
        for key in self.category_vars:
            self.category_vars[key].set("")
        self._hide_right_panel()

    def _select_day(self, day):
        # 같은 날짜 다시 클릭 시 패널 닫기
        if self.selected_day == day:
            self._hide_right_panel()
            return
        self.selected_day = day
        self._show_right_panel()
        self.selected_label.configure(
            text=f"{self.current_year}년 {self.current_month}월 {day}일",
            text_color=("gray10", "gray90"),
        )
        key = self._month_key()
        day_str = str(day)
        # 이 날짜의 행 개수 로드 (날짜마다 독립)
        saved = self.day_slots.get(key, {}).get(day_str)
        if saved and isinstance(saved, dict):
            self.slots_per_category = {c: max(1, int(saved.get(c, 1))) for c in CATEGORIES}
        else:
            self.slots_per_category = _default_slots().copy()
        self._build_category_rows()
        plan = self.plans.get(key, {})
        current = plan.get(day_str, "")
        if current:
            raw = [p.strip() for p in current.split(" | ")]
            n = len(CATEGORIES)
            # 구간이 4개 미만이면 앞(주식) 빈 칸 + 뒤(기타) 빈 칸이 빠진 것으로 복구
            if len(raw) < n:
                parts = ([""] + raw + [""] * max(0, n - len(raw) - 1))[:n]
            else:
                parts = raw[:n]
            for i, cat in enumerate(CATEGORIES):
                slot_vals = [s.strip() for s in parts[i].strip().split(",") if s.strip()] or [""]
                for slot in range(self.slots_per_category.get(cat, 1)):
                    k = (cat, slot)
                    if k in self.category_vars:
                        self.category_vars[k].set(slot_vals[slot] if slot < len(slot_vals) else "")
        else:
            for k in self.category_vars:
                self.category_vars[k].set("")

        for d, (btn, _) in self.day_buttons.items():
            if d == day:
                btn.configure(border_width=3, border_color="#1f6aa5")
            else:
                btn.configure(border_width=0, border_color=("gray85", "gray25"))

    def _apply_meal(self):
        """선택한 날짜에 현재 고른 메뉴를 반영(메모리만)."""
        if self.selected_day is None:
            return
        parts = []
        for cat in CATEGORIES:
            vals = []
            for slot in range(self.slots_per_category.get(cat, 1)):
                k = (cat, slot)
                if k in self.category_vars:
                    v = self.category_vars[k].get().strip()
                    if v:
                        vals.append(v)
            parts.append(",".join(vals))
        # 항상 주식|국/스프|반찬|기타 4구간으로 저장 (빈 칸도 유지)
        assert len(parts) == len(CATEGORIES), "parts must match categories"
        line = " | ".join(parts)
        key = self._month_key()
        if key not in self.plans:
            self.plans[key] = {}
        self.plans[key][str(self.selected_day)] = line
        btn, _ = self.day_buttons[self.selected_day]
        line_display = line.replace(" | ", "\n").replace(",", "\n") if line else ""
        btn.configure(text=f"{self.selected_day}\n{line_display}" if line_display else str(self.selected_day))

    def _save(self):
        """선택한 날 적용 후 data.json에 저장."""
        self._apply_meal()
        self._save_day_slots()
        save_all(self.menus, self.plans, self.day_slots)
        self.selected_label.configure(text="저장되었습니다.", text_color="green")
        self.after(1500, self._refresh_selected_label)

    def _refresh_selected_label(self):
        if self.selected_day:
            self.selected_label.configure(
                text=f"{self.current_year}년 {self.current_month}월 {self.selected_day}일",
                text_color=("gray10", "gray90"),
            )
        else:
            self.selected_label.configure(text="날짜를 클릭하세요", text_color="gray")

    def run(self):
        self.mainloop()


if __name__ == "__main__":
    app = MealPlannerApp()
    app.run()
